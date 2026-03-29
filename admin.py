"""
admin.py — SuperMart Admin Panel
Features: Product grid card view with real image upload support
Theme: Raspberry & Shades of Blue
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import os
import shutil
import database as db
import utils as u

# ── Images are stored in a folder next to the app ────────────────────────────
IMAGES_DIR = os.path.join(os.path.dirname(__file__), "product_images")
os.makedirs(IMAGES_DIR, exist_ok=True)


def _get_product_image_path(product_name):
    """Return the saved image path for a product, or None if not set."""
    safe_name = "".join(
        c if c.isalnum() or c in (" ", "_") else "_"
        for c in product_name
    ).strip().replace(" ", "_")
    for ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp"):
        path = os.path.join(IMAGES_DIR, safe_name + ext)
        if os.path.exists(path):
            return path
    return None


def _save_product_image(product_name, src_path):
    """Copy the chosen image file into product_images/ folder."""
    safe_name = "".join(
        c if c.isalnum() or c in (" ", "_") else "_"
        for c in product_name
    ).strip().replace(" ", "_")
    ext = os.path.splitext(src_path)[1].lower()
    # Remove old images for this product first
    for old_ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp"):
        old_path = os.path.join(IMAGES_DIR, safe_name + old_ext)
        if os.path.exists(old_path):
            os.remove(old_path)
    dest = os.path.join(IMAGES_DIR, safe_name + ext)
    shutil.copy2(src_path, dest)
    return dest


def _load_tk_image(path, size=(100, 90)):
    """Load an image using PIL if available, else return None."""
    try:
        from PIL import Image, ImageTk
        img = Image.open(path).resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None


class AdminPanel(tk.Tk):
    def __init__(self, user):
        super().__init__()
        self.user = user
        self._photo_refs = []   # keep PhotoImage references alive
        self.title(f"SuperMart — Admin Panel  [{user['username']}]")
        self.configure(bg=u.COLORS["bg"])
        u.apply_theme(self)
        
        # Natively Maximize the window
        try:
            self.state('zoomed') 
        except tk.TclError:
            w, h = self.winfo_screenwidth(), self.winfo_screenheight()
            self.geometry(f"{w}x{h}+0+0")

        self._build()
        self._load_products()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self):
        self._topbar()
        body = tk.Frame(self, bg=u.COLORS["bg"])
        body.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        nb = ttk.Notebook(body)
        nb.pack(fill="both", expand=True)

        self.tab_products = ttk.Frame(nb)
        self.tab_users    = ttk.Frame(nb)
        self.tab_reports  = ttk.Frame(nb)

        nb.add(self.tab_products, text="  📦  Products  ")
        nb.add(self.tab_users,    text="  👤  Users  ")
        nb.add(self.tab_reports,  text="  📊  Reports  ")

        self._build_products_tab()
        self._build_users_tab()
        self._build_reports_tab()

    def _topbar(self):
        C = u.COLORS
        bar = tk.Frame(self, bg=C["raspberry_dk"], height=54)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        left = tk.Frame(bar, bg=C["raspberry_dk"])
        left.pack(side="left", padx=16, fill="y")
        tk.Label(left, text="🛒", font=("Helvetica", 20),
                 bg=C["raspberry_dk"], fg="#fff").pack(side="left")
        tk.Label(left, text="  SuperMart", font=u.FONTS["heading"],
                 bg=C["raspberry_dk"], fg="#fff").pack(side="left")

        tk.Frame(bar, bg=C["raspberry"], width=2).pack(
            side="left", fill="y", pady=10, padx=10)
        tk.Label(bar, text="ADMIN PANEL", font=u.FONTS["sub"],
                 bg=C["raspberry_dk"], fg=C["blue_lt"]).pack(side="left")

        right = tk.Frame(bar, bg=C["raspberry_dk"])
        right.pack(side="right", padx=14, fill="y")
        tk.Label(right, text=f"👤 {self.user['username']}",
                 font=u.FONTS["small"], bg=C["raspberry_dk"],
                 fg=C["blue_lt"]).pack(side="left", padx=8)

        tk.Button(right, text="Open Billing", font=u.FONTS["btn"],
                  bg=C["blue"], fg="#fff", relief="flat", bd=0,
                  activebackground=C["blue2"], cursor="hand2",
                  padx=12, pady=4,
                  command=self._open_billing).pack(side="left", padx=4)
        tk.Button(right, text="Logout", font=u.FONTS["btn"],
                  bg=C["surface2"], fg=C["text"], relief="flat", bd=0,
                  activebackground=C["border"], cursor="hand2",
                  padx=12, pady=4,
                  command=self._logout).pack(side="left", padx=4)

    # ── Products Tab ──────────────────────────────────────────────────────────

    def _build_products_tab(self):
        C = u.COLORS
        tab = self.tab_products

        left = tk.Frame(tab, bg=C["surface"],
                        highlightbackground=C["raspberry_dk"],
                        highlightthickness=1)
        left.pack(side="left", fill="y", padx=(8, 4), pady=8)

        right = tk.Frame(tab, bg=C["bg"])
        right.pack(side="left", fill="both", expand=True,
                   padx=(4, 8), pady=8)

        self._build_product_form(left)
        self._build_product_grid(right)

    def _build_product_form(self, parent):
        C = u.COLORS

        # ── Fully Scrollable form area with Scrollbar ─────────────────────────
        container = tk.Frame(parent, bg=C["surface"])
        container.pack(fill="both", expand=True)

        # Increased width to 280 for better visibility
        form_canvas = tk.Canvas(container, bg=C["surface"],
                                highlightthickness=0, width=280)
        
        # Added Vertical Scrollbar
        vsb = ttk.Scrollbar(container, orient="vertical", command=form_canvas.yview)
        form_canvas.configure(yscrollcommand=vsb.set)

        vsb.pack(side="right", fill="y")
        form_canvas.pack(side="left", fill="both", expand=True)

        form_inner = tk.Frame(form_canvas, bg=C["surface"])
        form_window = form_canvas.create_window((0, 0), window=form_inner, anchor="nw")

        def _on_frame_configure(event):
            form_canvas.configure(scrollregion=form_canvas.bbox("all"))

        def _on_canvas_configure(event):
            form_canvas.itemconfig(form_window, width=event.width)

        form_inner.bind("<Configure>", _on_frame_configure)
        form_canvas.bind("<Configure>", _on_canvas_configure)

        # Mousewheel scrolling support
        def _on_mousewheel(event):
            form_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
        form_canvas.bind("<Enter>", lambda e: form_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        form_canvas.bind("<Leave>", lambda e: form_canvas.unbind_all("<MouseWheel>"))

        form = tk.Frame(form_inner, bg=C["surface"])
        form.pack(fill="x", expand=True, padx=16, pady=16)

        # ── Image preview box ─────────────────────────────────────────────────
        tk.Label(form, text="PRODUCT IMAGE", font=u.FONTS["sub"],
                 bg=C["surface"], fg=C["raspberry2"]).pack(
            anchor="w", pady=(0, 6))

        self.img_preview_frame = tk.Frame(
            form, bg=C["surface2"],
            width=180, height=120,
            highlightbackground=C["border"],
            highlightthickness=1)
        self.img_preview_frame.pack(pady=(0, 4))
        self.img_preview_frame.pack_propagate(False)

        self.img_preview_lbl = tk.Label(
            self.img_preview_frame,
            text="📦", font=("Helvetica", 44),
            bg=C["surface2"], fg=C["raspberry2"])
        self.img_preview_lbl.pack(expand=True)

        self._chosen_image_path = None

        img_btn_row = tk.Frame(form, bg=C["surface"])
        img_btn_row.pack(fill="x", pady=(0, 12))

        tk.Button(img_btn_row,
                  text="🖼  Upload Image",
                  font=u.FONTS["btn"],
                  bg=C["blue"], fg="#fff",
                  relief="flat", bd=0, cursor="hand2",
                  padx=8, pady=5,
                  activebackground=C["blue2"],
                  command=self._choose_image).pack(
            side="left", expand=True, fill="x", padx=(0, 2))

        tk.Button(img_btn_row,
                  text="✖ Remove",
                  font=u.FONTS["btn"],
                  bg=C["surface2"], fg=C["text_dim"],
                  relief="flat", bd=0, cursor="hand2",
                  padx=8, pady=5,
                  activebackground=C["border"],
                  command=self._remove_image).pack(
            side="left", padx=(2, 0))

        self._pil_hint = tk.Label(
            form,
            text="💡 Install Pillow for real images:\npip install Pillow",
            font=u.FONTS["small"],
            bg=C["surface"], fg=C["text_muted"],
            justify="left")
        self._pil_hint.pack(anchor="w", pady=(0, 8))
        self._check_pil_hint()

        # ── Product details ───────────────────────────────────────────────────
        tk.Frame(form, bg=C["border"], height=1).pack(fill="x", pady=(0, 10))

        tk.Label(form, text="PRODUCT DETAILS", font=u.FONTS["sub"],
                 bg=C["surface"], fg=C["raspberry2"]).pack(
            anchor="w", pady=(0, 10))

        self.edit_id    = None
        self.prod_name  = tk.StringVar()
        self.prod_price = tk.StringVar()
        self.prod_stock = tk.StringVar()
        self.prod_name.trace_add("write", self._update_icon_preview)

        for label, var, kw in [
            ("Product Name",  self.prod_name,  {}),
            ("Price (₹)",     self.prod_price, {}),
            ("Stock (units)", self.prod_stock, {}),
        ]:
            tk.Label(form, text=label, font=u.FONTS["label"],
                     bg=C["surface"], fg=C["text_dim"]).pack(
                anchor="w", pady=(6, 0))
            ttk.Entry(form, textvariable=var, width=26, **kw).pack(
                fill="x", ipady=4, pady=(2, 0))

        # ── Action buttons ────────────────────────────────────────────────────
        tk.Frame(form, bg=C["border"], height=1).pack(fill="x", pady=(12, 8))

        for text, bg, hover, cmd in [
            ("➕  Add Product",     C["raspberry"],  C["raspberry2"], self._add_product),
            ("✏️  Update Selected", C["blue"],       C["blue2"],      self._update_product),
            ("🗑  Delete Selected", "#B71C1C",       C["danger"],     self._delete_product),
            ("🔄  Clear Form",      C["surface2"],   C["border"],     self._clear_form),
        ]:
            tk.Button(form, text=text, font=u.FONTS["btn"],
                      bg=bg, fg="#fff", relief="flat", bd=0,
                      activebackground=hover, cursor="hand2",
                      pady=6, command=cmd).pack(fill="x", pady=2)

    # ── Image helpers ─────────────────────────────────────────────────────────

    def _check_pil_hint(self):
        try:
            from PIL import Image  # noqa
            self._pil_hint.config(
                text="✅ Pillow installed — real images supported!",
                fg=u.COLORS["success"])
        except ImportError:
            pass   

    def _choose_image(self):
        path = filedialog.askopenfilename(
            title="Choose Product Image",
            filetypes=[
                ("Image files", "*.png *.jpg *.jpeg *.bmp *.gif"),
                ("All files", "*.*"),
            ]
        )
        if not path:
            return

        self._chosen_image_path = path
        self._show_preview_image(path)
        u.show_toast(self, "Image selected! Click Add/Update to save.", kind="info")

    def _show_preview_image(self, path):
        photo = _load_tk_image(path, size=(176, 116))
        if photo:
            self._photo_refs.append(photo)
            self.img_preview_lbl.config(image=photo, text="")
        else:
            fname = os.path.basename(path)
            self.img_preview_lbl.config(
                text=f"🖼\n{fname[:18]}",
                image="",
                font=u.FONTS["small"],
                fg=u.COLORS["blue3"])

    def _remove_image(self):
        self._chosen_image_path = None
        name = self.prod_name.get().strip()
        if name and self.edit_id:
            existing = _get_product_image_path(name)
            if existing and messagebox.askyesno(
                    "Remove Image",
                    f"Remove the saved image for '{name}'?"):
                os.remove(existing)
                u.show_toast(self, "Image removed.", kind="warning")
        icon = u.get_product_icon(self.prod_name.get())
        self.img_preview_lbl.config(
            text=icon, image="",
            font=("Helvetica", 44),
            fg=u.COLORS["raspberry2"])

    def _update_icon_preview(self, *_):
        if self._chosen_image_path:
            return  
        name = self.prod_name.get()
        if self.edit_id:
            path = _get_product_image_path(name)
            if path:
                self._show_preview_image(path)
                return
        icon = u.get_product_icon(name)
        self.img_preview_lbl.config(
            text=icon, image="",
            font=("Helvetica", 44),
            fg=u.COLORS["raspberry2"])

    # ── Product Grid ──────────────────────────────────────────────────────────

    def _build_product_grid(self, parent):
        C = u.COLORS

        hdr = tk.Frame(parent, bg=C["bg"])
        hdr.pack(fill="x", pady=(0, 6))
        tk.Label(hdr, text="PRODUCT INVENTORY",
                 font=u.FONTS["sub"], bg=C["bg"],
                 fg=C["raspberry2"]).pack(side="left")
        tk.Frame(hdr, bg=C["raspberry_dk"], height=1).pack(
            side="left", fill="x", expand=True, padx=(10, 0), pady=7)

        canvas_frame = tk.Frame(parent, bg=C["bg"])
        canvas_frame.pack(fill="both", expand=True)

        self.grid_canvas = tk.Canvas(
            canvas_frame, bg=C["bg"],
            highlightthickness=0, bd=0)
        vsb = ttk.Scrollbar(canvas_frame, orient="vertical",
                             command=self.grid_canvas.yview)
        self.grid_canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.grid_canvas.pack(side="left", fill="both", expand=True)

        self.grid_inner = tk.Frame(self.grid_canvas, bg=C["bg"])
        self._grid_window = self.grid_canvas.create_window(
            (0, 0), window=self.grid_inner, anchor="nw")

        self.grid_inner.bind("<Configure>", self._on_grid_configure)
        self.grid_canvas.bind("<Configure>", self._on_canvas_configure)
        self.grid_canvas.bind("<MouseWheel>",
            lambda e: self.grid_canvas.yview_scroll(
                int(-1*(e.delta/120)), "units"))

    def _on_grid_configure(self, event):
        self.grid_canvas.configure(
            scrollregion=self.grid_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.grid_canvas.itemconfig(
            self._grid_window, width=event.width)

    def _load_products(self):
        products = db.get_all_products()
        self._photo_refs.clear()

        for widget in self.grid_inner.winfo_children():
            widget.destroy()

        COLS = 4
        for i, p in enumerate(products):
            self._make_product_card(
                self.grid_inner, p, i // COLS, i % COLS)

        for c in range(COLS):
            self.grid_inner.columnconfigure(c, weight=1)

    def _make_product_card(self, parent, product, row, col):
        C    = u.COLORS
        name = product["name"]
        icon = u.get_product_icon(name)
        img_path = _get_product_image_path(name)

        card = tk.Frame(parent, bg=C["surface"],
                        highlightbackground=C["border"],
                        highlightthickness=1,
                        cursor="hand2")
        card.grid(row=row, column=col, padx=6, pady=6, sticky="nsew")

        tk.Frame(card, bg=C["raspberry"], height=3).pack(fill="x")

        img_frame = tk.Frame(card, bg=C["surface2"],
                              width=140, height=95)
        img_frame.pack(fill="x", padx=8, pady=(8, 3))
        img_frame.pack_propagate(False)

        if img_path:
            photo = _load_tk_image(img_path, size=(138, 93))
            if photo:
                self._photo_refs.append(photo)
                tk.Label(img_frame, image=photo,
                         bg=C["surface2"]).pack(expand=True)
            else:
                tk.Label(img_frame, text="🖼",
                         font=("Helvetica", 38),
                         bg=C["surface2"],
                         fg=C["blue3"]).pack(expand=True)
        else:
            tk.Label(img_frame, text=icon,
                     font=("Helvetica", 40),
                     bg=C["surface2"],
                     fg=C["raspberry2"]).pack(expand=True)

        if img_path:
            badge = tk.Label(card, text="📷 Image",
                             font=u.FONTS["small"],
                             bg=C["blue"], fg="#fff",
                             padx=4, pady=1)
            badge.pack()
        else:
            tk.Label(card, text="No image",
                     font=u.FONTS["small"],
                     bg=C["surface"], fg=C["text_muted"]).pack()

        tk.Label(card, text=name,
                 font=u.FONTS["label_b"],
                 bg=C["surface"], fg=C["text"],
                 wraplength=140,
                 justify="center").pack(padx=6, pady=(3, 1))

        tk.Label(card, text=f"₹{float(product['price']):.2f}",
                 font=("Georgia", 13, "bold"),
                 bg=C["surface"],
                 fg=C["raspberry2"]).pack()

        stock_val   = product["stock"]
        stock_color = (C["danger"]  if stock_val <= 5  else
                       C["warning"] if stock_val <= 10 else
                       C["success"])
        stock_bg    = "#1A0A0A" if stock_val <= 5 else C["surface2"]

        sf = tk.Frame(card, bg=stock_bg,
                       highlightbackground=stock_color,
                       highlightthickness=1)
        sf.pack(padx=8, pady=(3, 8), fill="x")
        tk.Label(sf, text=f"Stock: {stock_val} units",
                 font=u.FONTS["small"],
                 bg=stock_bg, fg=stock_color).pack(pady=3)

        def on_click(event, p=product):
            self._select_product_from_card(p)

        self._bind_card_children(card, on_click)

        def on_enter(e, c=card):
            c.config(highlightbackground=C["raspberry2"],
                     highlightthickness=2)
        def on_leave(e, c=card):
            c.config(highlightbackground=C["border"],
                     highlightthickness=1)
        card.bind("<Enter>", on_enter)
        card.bind("<Leave>", on_leave)

    def _bind_card_children(self, widget, command):
        widget.bind("<Button-1>", command)
        for child in widget.winfo_children():
            self._bind_card_children(child, command)

    def _select_product_from_card(self, product):
        self.edit_id = product["id"]
        self.prod_name.set(product["name"])
        self.prod_price.set(str(float(product["price"])))
        self.prod_stock.set(str(product["stock"]))
        self._chosen_image_path = None

        path = _get_product_image_path(product["name"])
        if path:
            self._show_preview_image(path)
        else:
            icon = u.get_product_icon(product["name"])
            self.img_preview_lbl.config(
                text=icon, image="",
                font=("Helvetica", 44),
                fg=u.COLORS["raspberry2"])

        u.show_toast(self, f"Selected: {product['name']}", kind="info")

    # ── CRUD ──────────────────────────────────────────────────────────────────

    def _add_product(self):
        name  = self.prod_name.get().strip()
        price = self.prod_price.get().strip()
        stock = self.prod_stock.get().strip()

        if not all([name, price, stock]):
            messagebox.showwarning("Validation", "All fields are required.")
            return
        try:
            price = float(price)
            stock = int(stock)
            assert price > 0 and stock >= 0
        except (ValueError, AssertionError):
            messagebox.showerror("Validation",
                "Price must be positive; stock must be a whole number.")
            return

        ok, msg = db.add_product(name, price, stock)
        if ok:
            if self._chosen_image_path:
                try:
                    _save_product_image(name, self._chosen_image_path)
                except Exception as e:
                    messagebox.showwarning("Image", f"Product saved but image failed:\n{e}")
            self._load_products()
            self._clear_form()
            u.show_toast(self, msg, kind="success")
        else:
            messagebox.showerror("Error", msg)

    def _update_product(self):
        if self.edit_id is None:
            messagebox.showinfo("Selection",
                "Click on a product card first to select it.")
            return
        name  = self.prod_name.get().strip()
        price = self.prod_price.get().strip()
        stock = self.prod_stock.get().strip()

        if not all([name, price, stock]):
            messagebox.showwarning("Validation", "All fields are required.")
            return
        try:
            price = float(price)
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Validation", "Invalid price or stock value.")
            return

        ok, msg = db.update_product(self.edit_id, name, price, stock)
        if ok:
            if self._chosen_image_path:
                try:
                    _save_product_image(name, self._chosen_image_path)
                except Exception as e:
                    messagebox.showwarning("Image", f"Product updated but image failed:\n{e}")
            self._load_products()
            self._clear_form()
            u.show_toast(self, msg, kind="success")
        else:
            messagebox.showerror("Error", msg)

    def _delete_product(self):
        if self.edit_id is None:
            messagebox.showinfo("Selection",
                "Click on a product card first to select it.")
            return
        name = self.prod_name.get()
        if not messagebox.askyesno("Confirm",
                f"Delete '{name}'?"):
            return

        img_path = _get_product_image_path(name)
        if img_path and os.path.exists(img_path):
            os.remove(img_path)

        db.delete_product(self.edit_id)
        self._load_products()
        self._clear_form()
        u.show_toast(self, "Product deleted.", kind="warning")

    def _clear_form(self):
        self.edit_id = None
        self._chosen_image_path = None
        self.prod_name.set("")
        self.prod_price.set("")
        self.prod_stock.set("")
        self.img_preview_lbl.config(
            text="📦", image="",
            font=("Helvetica", 44),
            fg=u.COLORS["raspberry2"])

    # ── Users Tab ─────────────────────────────────────────────────────────────

    def _build_users_tab(self):
        C = u.COLORS
        tab = self.tab_users

        left = tk.Frame(tab, bg=C["surface"],
                        highlightbackground=C["raspberry_dk"],
                        highlightthickness=1)
        left.pack(side="left", fill="y", padx=(8, 4), pady=8)

        right = tk.Frame(tab, bg=C["bg"])
        right.pack(side="left", fill="both", expand=True,
                   padx=(4, 8), pady=8)

        form = tk.Frame(left, bg=C["surface"])
        form.pack(fill="x", padx=20, pady=20)

        tk.Label(form, text="ADD NEW USER", font=u.FONTS["sub"],
                 bg=C["surface"], fg=C["raspberry2"]).grid(
            row=0, column=0, sticky="w", pady=(0, 16))

        self.new_user = tk.StringVar()
        self.new_pass = tk.StringVar()
        self.new_role = tk.StringVar(value="user")

        for i, (lbl, var, kw) in enumerate([
            ("Username", self.new_user, {}),
            ("Password", self.new_pass, {"show": "●"}),
        ]):
            tk.Label(form, text=lbl, font=u.FONTS["label"],
                     bg=C["surface"], fg=C["text_dim"]).grid(
                row=i*2+1, column=0, sticky="w")
            ttk.Entry(form, textvariable=var, width=26, **kw).grid(
                row=i*2+2, column=0, sticky="ew",
                ipady=4, pady=(4, 10))

        tk.Label(form, text="Role", font=u.FONTS["label"],
                 bg=C["surface"], fg=C["text_dim"]).grid(
            row=5, column=0, sticky="w")
        ttk.Combobox(form, textvariable=self.new_role,
                     values=["user", "admin"],
                     state="readonly", width=24).grid(
            row=6, column=0, sticky="ew",
            ipady=4, pady=(4, 16))
        form.columnconfigure(0, weight=1)

        tk.Button(left, text="➕  Create User",
                  font=u.FONTS["btn"],
                  bg=C["raspberry"], fg="#fff",
                  relief="flat", bd=0, cursor="hand2",
                  pady=6, command=self._add_user).pack(
            fill="x", padx=20, pady=(4, 6))

        tk.Button(left, text="🗑  Delete Selected",
                  font=u.FONTS["btn"],
                  bg="#B71C1C", fg="#fff",
                  relief="flat", bd=0, cursor="hand2",
                  pady=6, command=self._delete_user).pack(
            fill="x", padx=20, pady=(0, 4))

        u.section_header(right, "System Users",
                         bg=C["bg"]).pack(fill="x", pady=(4, 6))

        cols  = ("id", "username", "role")
        heads = ("ID", "Username", "Role")
        tf, self.user_tree = u.scrollable_tree(
            right, cols, heads, height=20)
        tf.pack(fill="both", expand=True)
        self.user_tree.column("id",       width=50,  anchor="center")
        self.user_tree.column("username", width=200, anchor="w")
        self.user_tree.column("role",     width=100, anchor="center")
        self.user_tree.heading("id",       text="ID",       anchor="center")
        self.user_tree.heading("username", text="Username", anchor="w")
        self.user_tree.heading("role",     text="Role",     anchor="center")
        self._load_users()

    def _load_users(self):
        users = db.get_all_users()
        u.tree_reload(self.user_tree, users, ("id", "username", "role"))

    def _add_user(self):
        username = self.new_user.get().strip()
        password = self.new_pass.get().strip()
        role     = self.new_role.get()
        if not username or not password:
            messagebox.showwarning("Validation",
                "Username and password are required.")
            return
        ok = db.add_user(username, password, role)
        if ok:
            self.new_user.set("")
            self.new_pass.set("")
            self._load_users()
            u.show_toast(self, f"User '{username}' created.", kind="success")
        else:
            messagebox.showerror("Error", "Username already exists.")

    def _delete_user(self):
        selected_item = self.user_tree.selection()
        if not selected_item:
            messagebox.showinfo("Selection Required",
                "Please click on a user in the table first.")
            return
        item_values = self.user_tree.item(selected_item[0])["values"]
        user_id     = item_values[0]
        username    = item_values[1]
        if username == self.user["username"]:
            messagebox.showwarning("Action Denied",
                "You cannot delete the account you are logged into.")
            return
        if messagebox.askyesno("Confirm Deletion",
                f"Permanently delete user '{username}'?"):
            db.delete_user(user_id)
            self._load_users()
            u.show_toast(self, f"User '{username}' deleted.", kind="warning")

    # ── Reports Tab ───────────────────────────────────────────────────────────

    def _build_reports_tab(self):
        C = u.COLORS
        tab = self.tab_reports

        kpi_row = tk.Frame(tab, bg=C["bg"])
        kpi_row.pack(fill="x", padx=8, pady=12)

        self._kpi_cards = {}
        kpis = [
            ("Total Revenue",   "₹0", C["raspberry2"]),
            ("Today's Revenue", "₹0", C["blue3"]),
            ("Total Bills",     "0",  C["warning"]),
        ]
        for i, (title, val, color) in enumerate(kpis):
            c = tk.Frame(kpi_row, bg=C["surface"],
                         highlightbackground=C["raspberry_dk"],
                         highlightthickness=1)
            c.grid(row=0, column=i, padx=6, sticky="nsew")
            kpi_row.columnconfigure(i, weight=1)
            tk.Frame(c, bg=color, height=3).pack(fill="x")
            tk.Label(c, text=title, font=u.FONTS["label"],
                     bg=C["surface"],
                     fg=C["text_dim"]).pack(padx=16, pady=(12, 4))
            lbl = tk.Label(c, text=val,
                           font=("Georgia", 22, "bold"),
                           bg=C["surface"], fg=color)
            lbl.pack(padx=16, pady=(0, 12))
            self._kpi_cards[title] = lbl

        action_frame = tk.Frame(tab, bg=C["bg"])
        action_frame.pack(pady=4)
        ttk.Button(action_frame, text="🔄  Refresh Reports",
                   style="Ghost.TButton",
                   command=self._load_reports).pack(side="left", padx=5)
        ttk.Button(action_frame, text="📥  Export Data",
                   style="Ghost.TButton",
                   command=self._export_reports).pack(side="left", padx=5)

        row2 = tk.Frame(tab, bg=C["bg"])
        row2.pack(fill="both", expand=True, padx=8, pady=4)

        left2  = tk.Frame(row2, bg=C["bg"])
        left2.pack(side="left", fill="both", expand=True, padx=(0, 4))
        right2 = tk.Frame(row2, bg=C["bg"])
        right2.pack(side="left", fill="both", expand=True, padx=(4, 0))

        # UPDATED: Changed Monthly Revenue to a Detailed Sales History
        u.section_header(left2,  "Sales History (All Bills)", bg=C["bg"]).pack(fill="x", pady=(0, 6))
        u.section_header(right2, "Stock Alert (Low ≤ 10)", bg=C["bg"]).pack(fill="x", pady=(0, 6))

        # UPDATED: New Treeview columns for individual bills
        _, self.sales_tree = u.scrollable_tree(
            left2, ("bill_id", "date", "customer", "total"),
            ("Bill ID", "Date", "Customer", "Total (₹)"), height=12)
        self.sales_tree.master.pack(fill="both", expand=True)
        
        # Adjust column widths for the sales table
        self.sales_tree.column("bill_id", width=80, anchor="center")
        self.sales_tree.column("date", width=120, anchor="center")
        self.sales_tree.column("customer", width=150, anchor="w")
        self.sales_tree.column("total", width=100, anchor="e")

        _, self.stock_tree = u.scrollable_tree(
            right2, ("icon", "name", "stock", "price"),
            ("", "Product", "Stock", "Price"), height=12)
        self.stock_tree.master.pack(fill="both", expand=True)
        self.stock_tree.column("icon", width=40, anchor="center")

        self._load_reports()

    def _load_reports(self):
        total = db.get_total_revenue()
        today = db.get_today_revenue()
        sales = db.get_all_sales()

        self._kpi_cards["Total Revenue"].config(text=u.currency(total))
        self._kpi_cards["Today's Revenue"].config(text=u.currency(today))
        self._kpi_cards["Total Bills"].config(text=str(len(sales)))

        # UPDATED: Load individual bills into the new Sales History table
        for item in self.sales_tree.get_children():
            self.sales_tree.delete(item)
            
        for i, row in enumerate(sales):
            tag = "odd" if i % 2 else "even"
            # Extract data safely assuming standard database dictionary keys
            b_id = row.get("bill_id", row.get("id", f"B-{i+1}"))
            date = row.get("date", "N/A")
            cust = row.get("customer_name", row.get("customer", "Walk-in"))
            amt  = float(row.get("total", row.get("amount", 0.0)))
            
            self.sales_tree.insert("", "end", values=(
                b_id,
                date,
                cust,
                f"₹{amt:,.2f}"
            ), tags=(tag,))

        # Load Stock Alerts
        products = db.get_all_products()
        low = [p for p in products if p["stock"] <= 10]
        for item in self.stock_tree.get_children():
            self.stock_tree.delete(item)
            
        for i, p in enumerate(sorted(low, key=lambda x: x["stock"])):
            tag  = "odd" if i % 2 else "even"
            icon = u.get_product_icon(p["name"])
            self.stock_tree.insert("", "end", values=(
                icon, p["name"],
                p["stock"],
                f"₹{float(p['price']):.2f}",
            ), tags=(tag,))

    # ── Export ────────────────────────────────────────────────────────────────

    def _export_reports(self):
        file_types = [
            ("PDF Document",            "*.pdf"),    # NEW PDF Option
            ("Excel Document",          "*.xlsx"),
            ("CSV (Comma delimited)",   "*.csv"),
            ("Text Document",           "*.txt"),
        ]
        filepath = filedialog.asksaveasfilename(
            title="Export Reports",
            defaultextension=".pdf",
            filetypes=file_types,
            initialfile="SuperMart_Reports")
            
        if not filepath:
            return
            
        try:
            if filepath.endswith(".pdf"):
                self._export_to_pdf(filepath)
            elif filepath.endswith(".xlsx"):
                self._export_to_excel(filepath)
            elif filepath.endswith(".csv"):
                self._export_to_csv(filepath)
            elif filepath.endswith(".txt"):
                self._export_to_txt(filepath)
                
            messagebox.showinfo("Success", f"Reports exported successfully to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export file:\n{str(e)}")

    def _export_to_pdf(self, filepath):
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror("Missing Library", "Install fpdf first:\npip install fpdf")
            return

        pdf = FPDF()
        pdf.add_page()
        
        # Title
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "SuperMart System Reports", ln=True, align="C")
        pdf.ln(5)

        # KPI Summary
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "SUMMARY", ln=True)
        pdf.set_font("Arial", '', 11)
        for title, lbl in self._kpi_cards.items():
            # Replace ₹ with Rs. because standard PDF fonts lack the rupee symbol
            val = lbl.cget("text").replace("₹", "Rs. ") 
            pdf.cell(50, 8, f"{title}:", border=0)
            pdf.cell(50, 8, val, border=0, ln=True)
        pdf.ln(5)

        # Sales History
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "SALES HISTORY (ALL BILLS)", ln=True)
        pdf.set_font("Arial", 'B', 10)
        
        # Table Headers
        pdf.cell(35, 8, "Bill ID", border=1)
        pdf.cell(45, 8, "Date", border=1)
        pdf.cell(70, 8, "Customer", border=1)
        pdf.cell(35, 8, "Total", border=1, ln=True)
        
        # Table Rows
        pdf.set_font("Arial", '', 10)
        for child in self.sales_tree.get_children():
            vals = self.sales_tree.item(child)["values"]
            pdf.cell(35, 8, str(vals[0]), border=1)
            pdf.cell(45, 8, str(vals[1]), border=1)
            pdf.cell(70, 8, str(vals[2])[:35], border=1) # Truncate long names
            pdf.cell(35, 8, str(vals[3]).replace("₹", "Rs. "), border=1, ln=True)
        pdf.ln(5)

        # Stock Alerts
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "LOW STOCK ALERTS", ln=True)
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(100, 8, "Product Name", border=1)
        pdf.cell(40, 8, "Stock Remaining", border=1)
        pdf.cell(45, 8, "Price", border=1, ln=True)
        
        pdf.set_font("Arial", '', 10)
        for child in self.stock_tree.get_children():
            vals = self.stock_tree.item(child)["values"]
            pdf.cell(100, 8, str(vals[1])[:50], border=1)
            pdf.cell(40, 8, str(vals[2]), border=1)
            pdf.cell(45, 8, str(vals[3]).replace("₹", "Rs. "), border=1, ln=True)

        pdf.output(filepath)

    def _export_to_excel(self, filepath):
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            messagebox.showerror("Missing Library", "Install openpyxl first:\npip install openpyxl")
            return
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SuperMart Reports"
        bold = Font(bold=True)
        
        ws.append(["SUMMARY"])
        ws["A1"].font = bold
        r = 2
        for title, lbl in self._kpi_cards.items():
            ws.cell(row=r, column=1, value=title)
            ws.cell(row=r, column=2, value=lbl.cget("text"))
            r += 1
        r += 1
        
        ws.cell(row=r, column=1, value="SALES HISTORY").font = bold
        r += 1
        for col_num, h in enumerate(["Bill ID", "Date", "Customer", "Total"], 1):
            ws.cell(row=r, column=col_num, value=h).font = bold
        r += 1
        for child in self.sales_tree.get_children():
            for col_num, val in enumerate(self.sales_tree.item(child)["values"], 1):
                ws.cell(row=r, column=col_num, value=val)
            r += 1
        r += 1
        
        ws.cell(row=r, column=1, value="LOW STOCK ALERTS").font = bold
        r += 1
        for col_num, h in enumerate(["Product", "Stock", "Price"], 1):
            ws.cell(row=r, column=col_num, value=h).font = bold
        r += 1
        for child in self.stock_tree.get_children():
            vals = self.stock_tree.item(child)["values"]
            ws.cell(row=r, column=1, value=vals[1])
            ws.cell(row=r, column=2, value=vals[2])
            ws.cell(row=r, column=3, value=vals[3])
            r += 1
            
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        wb.save(filepath)

    def _export_to_csv(self, filepath):
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["SUMMARY"])
            for title, lbl in self._kpi_cards.items():
                w.writerow([title, lbl.cget("text")])
            w.writerow([])
            
            w.writerow(["SALES HISTORY"])
            w.writerow(["Bill ID", "Date", "Customer", "Total"])
            for child in self.sales_tree.get_children():
                w.writerow(self.sales_tree.item(child)["values"])
            w.writerow([])
            
            w.writerow(["LOW STOCK ALERTS"])
            w.writerow(["Product", "Stock", "Price"])
            for child in self.stock_tree.get_children():
                vals = self.stock_tree.item(child)["values"]
                w.writerow([vals[1], vals[2], vals[3]])

    def _export_to_txt(self, filepath):
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("=== SUPERMART SYSTEM REPORTS ===\n\n")
            f.write("--- SUMMARY ---\n")
            for title, lbl in self._kpi_cards.items():
                f.write(f"{title}: {lbl.cget('text')}\n")
                
            f.write("\n--- SALES HISTORY ---\n")
            f.write(f"{'Bill ID':<15} | {'Date':<20} | {'Customer':<20} | {'Total':<10}\n")
            f.write("-" * 75 + "\n")
            for child in self.sales_tree.get_children():
                vals = self.sales_tree.item(child)["values"]
                f.write(f"{str(vals[0]):<15} | {str(vals[1]):<20} | {str(vals[2]):<20} | {str(vals[3]):<10}\n")
                
            f.write("\n--- LOW STOCK ALERTS ---\n")
            f.write(f"{'Product':<25} | {'Stock':<10} | {'Price':<15}\n")
            f.write("-" * 55 + "\n")
            for child in self.stock_tree.get_children():
                vals = self.stock_tree.item(child)["values"]
                f.write(f"{str(vals[1]):<25} | {str(vals[2]):<10} | {str(vals[3]):<15}\n")

    # ── Navigation ────────────────────────────────────────────────────────────

    def _open_billing(self):
        self.destroy()
        from billing import BillingScreen
        BillingScreen(self.user).mainloop()

    def _logout(self):
        self.destroy()
        from login import LoginScreen
        LoginScreen().mainloop()